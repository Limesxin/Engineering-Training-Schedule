import streamlit as st
import pandas as pd
import re
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import json
import os
import io

# ==========================================
# 1. 网页全局与核心参数配置
# ==========================================
st.set_page_config(page_title="工程训练排课系统", page_icon="🛠️", layout="wide")

ADMIN_PASSWORD = "888"
MASTER_FILE = '2025-2026工程训练_信息全满终极版.xlsx'
SUB_FILE = '各工种场地课表_最新版.xlsx'
CONFIG_FILE = 'custom_configs.json'

GLOBAL_PATTERN = re.compile(r'^([AB]?)\s*(.*?)\s*(\d+\'?-\d+\'?|考\d+)\s*(?:[（\(](.*?)[）\)])?\s*$')


# ==========================================
# 2. 核心功能一：读取与缓存数据
# ==========================================
@st.cache_data
def load_all_data():
    df_master = pd.read_excel(MASTER_FILE, sheet_name='排课表')
    df_master = df_master.fillna("")

    all_sub_sheets = pd.read_excel(SUB_FILE, sheet_name=None)
    for key in all_sub_sheets:
        all_sub_sheets[key] = all_sub_sheets[key].fillna("")

    return df_master, all_sub_sheets


def load_configs():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def save_configs(configs):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(configs, f, ensure_ascii=False, indent=4)


# ==========================================
# 3. 核心生成器：动态组装专属课表与 Excel 导出
# ==========================================
def generate_custom_df(df_master, selected_ws, selected_teachers):
    week_cols = {col: int(re.search(r'第(\d+)周', col).group(1)) for col in df_master.columns if
                 re.search(r'第(\d+)周', col)}
    days = ['周一', '周二', '周三', '周四', '周五']
    custom_schedule = {w: {d: {'上午': [], '下午': []} for d in days} for w in range(1, 22)}

    for index, row in df_master.iterrows():
        class_name = str(row.get('教学班名称', '')).strip()
        day = str(row.get('星期', '')).strip()
        if day not in days or not class_name: continue

        for col_name, week_num in week_cols.items():
            val = str(row[col_name])
            if val and val != 'nan':
                for line in val.split('\n'):
                    line = line.strip()
                    if not line: continue
                    match = GLOBAL_PATTERN.match(line)
                    if match:
                        ws_name = match.group(2).strip()
                        time_suffix = match.group(3)
                        teacher_name = match.group(4).strip() if match.group(4) else ""
                        if not ws_name: ws_name = '考试' if '考' in time_suffix else '未命名项目'

                        if (ws_name in selected_ws) or (teacher_name in selected_teachers):
                            display_text = f"{class_name} {line}"
                            is_am, is_pm = False, False
                            if '考' in time_suffix:
                                num = int(time_suffix.replace('考', ''))
                                if num <= 4: is_am = True
                                if num >= 5: is_pm = True
                            else:
                                parts = time_suffix.split('-')
                                start = int(parts[0].replace("'", ""))
                                end = int(parts[1].replace("'", ""))
                                if start <= 4: is_am = True
                                if end >= 5: is_pm = True

                            if is_am: custom_schedule[week_num][day]['上午'].append(display_text)
                            if is_pm: custom_schedule[week_num][day]['下午'].append(display_text)

    custom_data = []
    for w in range(1, 22):
        row_dict = {'周次': f'第{w}周'}
        for d in days:
            am_list = custom_schedule[w][d]['上午']
            pm_list = custom_schedule[w][d]['下午']
            am_str = "【上午】\n" + "\n".join(am_list) if am_list else "【上午】"
            pm_str = "【下午】\n" + "\n".join(pm_list) if pm_list else "【下午】"
            row_dict[d] = "" if not am_list and not pm_list else f"{am_str}\n{'-' * 18}\n{pm_str}"
        custom_data.append(row_dict)

    return pd.DataFrame(custom_data)


def to_excel_bytes(df, sheet_title="专属课表"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_title)
        worksheet = writer.sheets[sheet_title]
        worksheet.column_dimensions['A'].width = 12
        for col_idx in range(2, 7):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 32
        fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                if cell.row == 1:
                    cell.font = Font(bold=True)
                elif cell.row % 2 == 0:
                    cell.fill = fill_gray
    return output.getvalue()


# ==========================================
# 4. 美化渲染与同步引擎
# ==========================================
def display_multiline_table(df, freeze_mode="智能自适应 (推荐)"):
    cols = list(df.columns)
    col_class = '教学班名称'
    col_day = '星期'
    is_master_table = (col_class in cols and col_day in cols)

    if is_master_table:
        cols.remove(col_class)
        cols.remove(col_day)
        df = df[[col_class, col_day] + cols]

    df_display = df.replace(r'\n', '<br>', regex=True)
    html = df_display.to_html(escape=False, index=False)

    css = """<style>
.table-wrapper { max-height: 75vh; overflow-x: auto; overflow-y: auto; border: 1px solid #e6e9ef; border-radius: 5px; }
.custom-excel-table table { width: 100%; border-collapse: collapse; font-size: 14px; font-family: sans-serif; }
.custom-excel-table th { background-color: #f0f2f6; color: #31333F; border: 1px solid #e6e9ef; padding: 10px; text-align: center !important; white-space: nowrap; position: sticky; top: 0; z-index: 2; }
.custom-excel-table td { border: 1px solid #e6e9ef; padding: 8px; text-align: center; vertical-align: middle; line-height: 1.6; }
"""
    if is_master_table:
        freeze_css = ".custom-excel-table th:nth-child(1), .custom-excel-table td:nth-child(1) { position: sticky; left: 0; min-width: 220px; max-width: 220px; background-color: #ffffff; z-index: 1; }\n.custom-excel-table th:nth-child(2), .custom-excel-table td:nth-child(2) { position: sticky; left: 220px; min-width: 60px; max-width: 60px; background-color: #ffffff; z-index: 1; }\n.custom-excel-table th:nth-child(1), .custom-excel-table th:nth-child(2) { background-color: #e2e6f0; z-index: 3; }"
    else:
        freeze_css = ".custom-excel-table th:nth-child(1), .custom-excel-table td:nth-child(1) { position: sticky; left: 0; min-width: 80px; max-width: 80px; background-color: #ffffff; z-index: 1; }\n.custom-excel-table th:nth-child(1) { background-color: #e2e6f0; z-index: 3; }"

    mobile_opts = "@media screen and (max-width: 768px) { .custom-excel-table table { font-size: 12px; } .custom-excel-table th, .custom-excel-table td { padding: 6px 4px; } }"
    if freeze_mode == "智能自适应 (推荐)":
        css += f"@media screen and (min-width: 769px) {{\n{freeze_css}\n}}\n{mobile_opts}"
    elif freeze_mode == "🔒 强制冻结":
        css += f"{freeze_css}\n{mobile_opts}"
    elif freeze_mode == "🔓 取消冻结":
        css += f"{mobile_opts}"
    css += "</style>"

    st.write(css + f'<div class="table-wrapper custom-excel-table">{html}</div>', unsafe_allow_html=True)


def sync_sub_sheets(df_master_latest):
    week_cols = {col: int(re.search(r'第(\d+)周', col).group(1)) for col in df_master_latest.columns if
                 re.search(r'第(\d+)周', col)}
    workshop_schedule = {}
    days = ['周一', '周二', '周三', '周四', '周五']

    def init_ws(ws_name):
        if ws_name not in workshop_schedule:
            workshop_schedule[ws_name] = {w: {d: {'上午': [], '下午': []} for d in days} for w in range(1, 22)}

    for index, row in df_master_latest.iterrows():
        class_name = str(row.get('教学班名称', '')).strip()
        day = str(row.get('星期', '')).strip()
        if day not in days or not class_name: continue
        for col_name, week_num in week_cols.items():
            val = str(row[col_name])
            if val and val != 'nan':
                for line in val.split('\n'):
                    line = line.strip()
                    if not line: continue
                    match = GLOBAL_PATTERN.match(line)
                    if match:
                        ws_name = match.group(2).strip()
                        time_suffix = match.group(3)
                        if not ws_name: ws_name = '考试' if '考' in time_suffix else '未命名项目'
                        init_ws(ws_name)
                        display_text = f"{class_name} {line}"
                        is_am, is_pm = False, False
                        if '考' in time_suffix:
                            num = int(time_suffix.replace('考', ''))
                            if num <= 4: is_am = True
                            if num >= 5: is_pm = True
                        else:
                            parts = time_suffix.split('-')
                            start = int(parts[0].replace("'", ""))
                            end = int(parts[1].replace("'", ""))
                            if start <= 4: is_am = True
                            if end >= 5: is_pm = True
                        if is_am: workshop_schedule[ws_name][week_num][day]['上午'].append(display_text)
                        if is_pm: workshop_schedule[ws_name][week_num][day]['下午'].append(display_text)
    with pd.ExcelWriter(SUB_FILE, engine='openpyxl') as writer:
        all_ws = list(workshop_schedule.keys())
        if '理论' in all_ws:
            all_ws.remove('理论')
            all_ws = ['理论'] + sorted(all_ws)
        else:
            all_ws = sorted(all_ws)
        for ws in all_ws:
            ws_data = []
            for w in range(1, 22):
                row_dict = {'周次': f'第{w}周'}
                for d in days:
                    am_list = workshop_schedule[ws][w][d]['上午']
                    pm_list = workshop_schedule[ws][w][d]['下午']
                    am_str = "【上午】\n" + "\n".join(am_list) if am_list else "【上午】"
                    pm_str = "【下午】\n" + "\n".join(pm_list) if pm_list else "【下午】"
                    row_dict[d] = "" if not am_list and not pm_list else f"{am_str}\n{'-' * 18}\n{pm_str}"
                ws_data.append(row_dict)
            df_ws = pd.DataFrame(ws_data)
            df_ws.to_excel(writer, index=False, sheet_name=ws)
            worksheet = writer.sheets[ws]
            worksheet.column_dimensions['A'].width = 12
            for col_idx in range(2, 7): worksheet.column_dimensions[get_column_letter(col_idx)].width = 32
            fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                           max_col=worksheet.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    if cell.row == 1:
                        cell.font = Font(bold=True)
                    elif cell.row % 2 == 0:
                        cell.fill = fill_gray


def push_to_github(file_path, commit_message):
    try:
        from github import Github
        g = Github(st.secrets["GITHUB_TOKEN"])
        # ⚠️⚠️⚠️ 请务必修改这里！填入您的 GitHub 用户名和仓库名
        repo = g.get_repo("Limesxin/Engineering-Training-Schedule")
        with open(file_path, "rb") as f:
            content = f.read()
        try:
            file_online = repo.get_contents(file_path)
            repo.update_file(file_online.path, commit_message, content, file_online.sha)
        except:
            repo.create_file(file_path, commit_message, content)
    except Exception as e:
        st.error(f"GitHub 同步异常: {e}")


# ==========================================
# 5. 网页前端 UI 与交互逻辑
# ==========================================
st.title("🛠️ 工程训练排课与场地查询系统")

try:
    df_master, all_sub_sheets = load_all_data()
    saved_configs = load_configs()
except FileNotFoundError as e:
    st.error(f"❌ 找不到文件，请确保总表和分表都在同一个文件夹内！")
    st.stop()

# --- 侧边栏导航 ---
st.sidebar.header("⚙️ 导航与控制面板")
view_mode = st.sidebar.radio("👀 请选择视图模式：", [
    "📚 查看大总表",
    "📍 查看场地分表",
    "🔎 专属课表快速查询",
    "🧑‍🏫 个人专属课表 (新建与配置)"
])
st.sidebar.markdown("---")

st.sidebar.subheader("🎛️ 视图显示设置")
freeze_option = st.sidebar.radio("左侧列冻结模式：", ["智能自适应 (推荐)", "🔒 强制冻结", "🔓 取消冻结"])
st.sidebar.markdown("---")

st.sidebar.subheader("🔒 管理员通道")
input_pwd = st.sidebar.text_input("请输入修改密码解锁编辑模式：", type="password")
is_admin = (input_pwd == ADMIN_PASSWORD)
if is_admin: st.sidebar.success("✅ 密码正确，已解锁！")

# --- 【新增】下载全局双表按钮 ---
st.sidebar.markdown("---")
st.sidebar.subheader("📥 核心数据导出备份")
try:
    with open(MASTER_FILE, "rb") as f:
        master_bytes = f.read()
    with open(SUB_FILE, "rb") as f:
        sub_bytes = f.read()
    st.sidebar.download_button("📦 下载最新【全景大总表】", data=master_bytes, file_name="最新_工程训练总表.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
    st.sidebar.download_button("📦 下载最新【各工种场地表】", data=sub_bytes, file_name="最新_各工种场地课表.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
except Exception:
    pass  # 若文件正在写入可能偶发读取失败，静默处理

# ----------------- 模式一：大总表 -----------------
if view_mode == "📚 查看大总表":
    st.subheader("📌 当前视图：【工程训练全景大总表】")
    if is_admin:
        st.info("💡 提示：修改完成后，请点击下方保存按钮。")
        edited_df = st.data_editor(df_master, use_container_width=True, hide_index=True, num_rows="dynamic")
        if st.button("💾 保存修改，并永久同步至 GitHub", type="primary"):
            with st.spinner("🚀 正在保存并向 GitHub 数据库进行永久同步，请千万不要关闭网页（约需10秒）..."):
                edited_df.to_excel(MASTER_FILE, index=False, sheet_name='排课表')
                sync_sub_sheets(edited_df)
                push_to_github(MASTER_FILE, "Web App Auto Sync Master")
                push_to_github(SUB_FILE, "Web App Auto Sync Sub Sheets")
                st.cache_data.clear()
            st.success("🎉 修改已成功保存并永久同步！页面即将刷新...")
            import time;

            time.sleep(2);
            st.rerun()
    else:
        display_multiline_table(df_master, freeze_option)

# ----------------- 模式二：单一场地分表 -----------------
elif view_mode == "📍 查看场地分表":
    sheet_names = list(all_sub_sheets.keys())
    selected_sheet = st.sidebar.selectbox("🎯 请选择要查看的场地/工种：", sheet_names)
    st.subheader(f"📌 当前视图：【{selected_sheet}】场地课表")
    display_multiline_table(all_sub_sheets[selected_sheet], freeze_option)

# ----------------- 模式三：专属课表快速查询 (带管理员删除) -----------------
elif view_mode == "🔎 专属课表快速查询":
    st.subheader("📌 当前视图：【专属课表快速查询】")
    if not saved_configs:
        st.warning("📭 目前还没有人保存过专属课表。请先去『个人专属课表 (新建与配置)』里创建一个吧！")
    else:
        config_names = list(saved_configs.keys())
        selected_config_name = st.selectbox("🎯 请选择已保存的专属课表：", config_names)

        cfg = saved_configs[selected_config_name]
        df_custom = generate_custom_df(df_master, cfg['ws'], cfg['teachers'])

        excel_bytes = to_excel_bytes(df_custom, sheet_title=selected_config_name)
        st.download_button(
            label=f"📥 一键下载【{selected_config_name}】 (Excel格式)",
            data=excel_bytes,
            file_name=f"{selected_config_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

        # 【新增：管理员删除功能】
        if is_admin:
            st.markdown("---")
            st.error("🚨 管理员危险操作区：清理废弃课表")
            col_del1, col_del2 = st.columns([3, 1])
            with col_del1:
                del_target = st.selectbox("请选择要永久删除的废弃课表：", config_names, key="del_select")
            with col_del2:
                st.write("")
                if st.button("🗑️ 确认永久删除", use_container_width=True):
                    with st.spinner("正在从云端数据库中抹除该配置..."):
                        del saved_configs[del_target]
                        save_configs(saved_configs)
                        push_to_github(CONFIG_FILE, f"Delete custom schedule: {del_target}")
                    st.success(f"已成功删除【{del_target}】！页面即将刷新...")
                    import time;

                    time.sleep(1.5);
                    st.rerun()

        st.markdown(
            f"**当前配置详情**：涵盖工种 `[ {', '.join(cfg['ws'])} ]` ，代课标识 `[ {', '.join(cfg['teachers'])} ]`")
        display_multiline_table(df_custom, freeze_option)

# ----------------- 模式四：个人专属组合与新建 -----------------
elif view_mode == "🧑‍🏫 个人专属课表 (新建与配置)":
    st.subheader("📌 当前视图：【个人专属自由组合与配置】")

    available_ws, available_teachers = set(), set()
    week_cols = {col: int(re.search(r'第(\d+)周', col).group(1)) for col in df_master.columns if
                 re.search(r'第(\d+)周', col)}

    for index, row in df_master.iterrows():
        for col in week_cols.keys():
            val = str(row[col])
            if val and val != 'nan':
                for line in val.split('\n'):
                    match = GLOBAL_PATTERN.match(line.strip())
                    if match:
                        ws = match.group(2).strip()
                        teacher = match.group(4).strip() if match.group(4) else ""
                        if not ws: ws = '考试' if '考' in match.group(3) else '未命名项目'
                        available_ws.add(ws)
                        if teacher: available_teachers.add(teacher)

    col1, col2 = st.columns(2)
    with col1:
        selected_ws = st.multiselect("🎯 1. 请选择工种 (可多选):", sorted(list(available_ws)))
    with col2:
        selected_teachers = st.multiselect("👤 2. 请选择代课标识 (可多选):", sorted(list(available_teachers)))

    if not selected_ws and not selected_teachers:
        st.warning("👈 请在上方至少选择一项工种或代课标识以预览您的专属课表。")
    else:
        df_custom = generate_custom_df(df_master, selected_ws, selected_teachers)
        st.success("✨ 预览生成完毕！如果不保存，此表在离开页面后会消失。")
        display_multiline_table(df_custom, freeze_option)

        st.markdown("---")
        st.subheader("💾 保存到云端快速查询库")
        col_input, col_btn = st.columns([3, 1])
        with col_input:
            remark_name = st.text_input("请给这个课表起个名字（例：张三老师的周课表）：", placeholder="输入备注名称...")
        with col_btn:
            st.write("")
            if st.button("🚀 保存并公开此专属课表", use_container_width=True):
                if not remark_name.strip():
                    st.error("❌ 名字不能为空哦！")
                else:
                    with st.spinner("正在写入云端配置库..."):
                        saved_configs[remark_name.strip()] = {
                            "ws": selected_ws,
                            "teachers": selected_teachers
                        }
                        save_configs(saved_configs)
                        push_to_github(CONFIG_FILE, f"Add new custom schedule: {remark_name}")
                    st.success(f"🎉 保存成功！您现在可以去左侧导航栏的『🔎 专属课表快速查询』里直接下拉调用它啦！")